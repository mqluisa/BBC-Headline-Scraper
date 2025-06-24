import openpyxl
import requests
from bs4 import BeautifulSoup
from datetime import datetime

# Para extrair a manchete e a data de publicação
def extrair_info(url):
    try:
        # Checar se a URL é válida
        if not url or not url.startswith('http'):
            return "URL inválida", "URL inválida"
        
        response = requests.get(url)
        response.raise_for_status()

        # Parsing do conteúdo com BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')

        # Extração da manchete (a tag <h1>)
        manchete = soup.find('h1')
        texto_manchete = manchete.get_text(strip=True) if manchete else "Não encontrada"
        
        # Extração da data de publicação da tag <time> (com atributo 'datetime')
        data_pub_tag = soup.find('time')
        data_pub = data_pub_tag['datetime'] if data_pub_tag else None

        # Converter a data para a ordem dia, mês e ano
        if data_pub:
            try:
                data_pub = datetime.fromisoformat(data_pub.replace("Z", "+00:00")).strftime("%d/%m/%Y")
            except Exception as e:
                print(f"Erro no parsing da data: {e}")
                data_pub = "Data inválida"
        else:
            data_pub = "Data de Publicação não encontrada"

        return texto_manchete, data_pub

    except Exception as e:
        print(f"Erro de processamento da URL {url}: {e}")
        return "Erro", "Erro"

# Carregar workbook e selecionar sheet
try:
    wb = openpyxl.load_workbook(' ')  # Inserir entre os parênteses o nome e localização do arquivo desejado
    sheet = wb.active
except FileNotFoundError:
    print("Erro: O arquivo desejado não foi encontrado.")
    exit()

coluna_link = None
coluna_manchete = None
coluna_data_pub = None

# Encontrar as colunas
for num_coluna, cell in enumerate(sheet[1], 1):  # A primeira linha contém os cabeçalhos
    if cell.value == "Link":
        coluna_link = num_coluna
    elif cell.value == "Manchete":
        coluna_manchete = num_coluna
    elif cell.value == "Data de Publicação":
        coluna_data_pub = num_coluna

if coluna_link is None or coluna_manchete is None or coluna_data_pub is None:
    print("Erro: Coluna de 'Link', 'Manchete', ou 'Data de Publicação' não encontrada")
    exit()

# Começar pela segunda linha (pular o cabeçalho)
for num_linha in range(2, sheet.max_row + 1):
    url = sheet.cell(row=num_linha, column=coluna_link).value
    print(f"Linha {num_linha}: Processando URL: {url}")
    if url:
        manchete, data_pub = extrair_info(url)
        sheet.cell(row=num_linha, column=coluna_manchete).value = manchete
        sheet.cell(row=num_linha, column=coluna_data_pub).value = data_pub
    else:
        print(f"Linha {num_linha}: Nenhuma URL encontrada. Pulando linha.")

# Salvar as informações atualizadas
wb.save(' ') # Inserir entre os parênteses o nome e localização do arquivo desejado
print("As manchetes e datas de publicação foram atualizadas.")
